"""Test del exportador Excel: hojas, formatos y precios numéricos."""
from openpyxl import load_workbook

from scraper.excel_export import export_to_excel
from scraper.models import Product


def _product(**overrides):
    base = dict(
        cadena="Test", nombre="Aceite de oliva", marca="Marca",
        categoria_nativa="Despensa > Aceites", categoria="Despensa",
        precio=3.90, precio_unidad=3.90, unidad="L", ean="8412345678901",
        url="https://example.com/p/1", disponible=True, promocion=None,
    )
    base.update(overrides)
    return Product(**base)


def test_export(tmp_path):
    products = [
        _product(),
        _product(nombre="Leche entera", categoria="Lácteos y Huevos", precio=0.98),
        _product(nombre="Sin precio", categoria="Lácteos y Huevos", precio=None),
    ]
    path = export_to_excel(products, out_dir=tmp_path, filename="test.xlsx")
    wb = load_workbook(path)

    assert wb.sheetnames == ["Todos", "Despensa", "Lácteos y Huevos"]
    ws = wb["Todos"]
    assert ws.freeze_panes == "A2"
    assert ws.tables  # autofiltro vía tabla
    assert ws.max_row == 4

    # precios como número con formato de moneda, no texto
    price_cell = ws.cell(row=2, column=6)
    assert isinstance(price_cell.value, float)
    assert "€" in price_cell.number_format

    # hoja por categoría con solo sus productos
    assert wb["Lácteos y Huevos"].max_row == 3


def test_export_empty(tmp_path):
    path = export_to_excel([], out_dir=tmp_path, filename="empty.xlsx")
    wb = load_workbook(path)
    assert wb.sheetnames == ["Todos"]


def test_sheet_name_sanitized(tmp_path):
    products = [_product(categoria="Una categoría con nombre larguísimo / inválido: [sí]")]
    path = export_to_excel(products, out_dir=tmp_path, filename="s.xlsx")
    wb = load_workbook(path)
    assert len(wb.sheetnames) == 2
    assert all(len(s) <= 31 for s in wb.sheetnames)
