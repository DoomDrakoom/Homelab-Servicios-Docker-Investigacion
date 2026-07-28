"""Exportación de productos a Excel con openpyxl.

- Hoja "Todos" con todos los productos, tabla con autofiltro y cabecera congelada.
- Una hoja adicional por categoría unificada.
- Los precios se escriben como número con formato de moneda, nunca como texto.
"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .models import Product

COLUMNS = [
    ("Cadena", "cadena", 14),
    ("Nombre", "nombre", 52),
    ("Marca", "marca", 18),
    ("Categoría", "categoria", 22),
    ("Categoría nativa", "categoria_nativa", 42),
    ("Precio (€)", "precio", 11),
    ("Precio/unidad (€)", "precio_unidad", 15),
    ("Unidad", "unidad", 9),
    ("EAN", "ean", 15),
    ("Disponible", "disponible", 11),
    ("Promoción", "promocion", 30),
    ("URL", "url", 45),
    ("Fecha extracción", "timestamp", 20),
]

CURRENCY_FORMAT = '#,##0.00 "€"'
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _sheet_name(name: str, used: set[str]) -> str:
    clean = _INVALID_SHEET_CHARS.sub("", name)[:31] or "Hoja"
    candidate, n = clean, 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _table_name(sheet_name: str, used: set[str]) -> str:
    base = "t_" + re.sub(r"\W", "_", sheet_name)
    candidate, n = base, 2
    while candidate.lower() in used:
        candidate = f"{base}_{n}"
        n += 1
    used.add(candidate.lower())
    return candidate


def _write_sheet(ws: Worksheet, products: Iterable[Product], table_names: set[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="217346")
    for col, (title, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 1
    for row, product in enumerate(products, start=2):
        for col, (_, attr, _w) in enumerate(COLUMNS, start=1):
            value = getattr(product, attr)
            if attr == "disponible":
                value = "Sí" if value else "No"
            cell = ws.cell(row=row, column=col, value=value)
            if attr in ("precio", "precio_unidad") and value is not None:
                cell.number_format = CURRENCY_FORMAT

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ref = f"A1:{last_col}{max(row, 2)}"
    table = Table(displayName=_table_name(ws.title, table_names), ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)


def export_to_excel(products: list[Product], out_dir: str | Path = "data/exports",
                    filename: str | None = None) -> Path:
    """Genera el .xlsx y devuelve su ruta."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = filename or f"productos_{date.today().strftime('%Y%m%d')}.xlsx"
    path = out_dir / filename

    workbook = Workbook()
    sheet_names: set[str] = set()
    table_names: set[str] = set()

    master = workbook.active
    master.title = _sheet_name("Todos", sheet_names)
    _write_sheet(master, products, table_names)

    by_category: dict[str, list[Product]] = {}
    for product in products:
        by_category.setdefault(product.categoria or "Otros", []).append(product)
    for category in sorted(by_category):
        ws = workbook.create_sheet(_sheet_name(category, sheet_names))
        _write_sheet(ws, by_category[category], table_names)

    workbook.save(path)
    return path
